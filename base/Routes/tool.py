import pywhatkit as kit
from django.conf import settings
from django.shortcuts import render, redirect
from googletrans import Translator, LANGUAGES
from django.http import HttpResponse
from gtts import gTTS
from langdetect import detect
import os
import io
from LMS.settings import BASE_DIR
import wikipedia
from docx2pdf import convert
from django.core.files.storage import default_storage
from pdf2docx import parse
import tabula
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from PIL import Image
import tempfile
import requests
from random import choice
from docx import Document
from django.http import JsonResponse
from docx.shared import Inches
from googlesearch import search
from bs4 import BeautifulSoup
from .Tool.Code_scriping_Tool import get_image_url
from .Tool.Tools import student_detials, staff_detials
import random
import nltk
from nltk.corpus import wordnet

from .Tool.Code_scriping_Tool import get_stackoverflow_link, get_stackoverflow_link_1, get_example_code_gfg, get_answer_from_given_link


def toolHome(request):
    return render(request, "tools/ToolHome.html", student_detials(request, 'Tools-Home'))


def Code_scriping(request):
    context = {}
    if request.method == 'POST':
        question = request.POST.get('question')
        if question:
            # Get the Stack Overflow link for the question
            link = get_stackoverflow_link_1(question)
            link_gfg = get_stackoverflow_link_1(question, 'geeksforgeeks.org')
            if link:
                # Get the example code from the link
                code = get_answer_from_given_link(link)
                code_gfg = get_example_code_gfg(link_gfg)
                if code:
                    # Add the question, link and code to the context
                    context['question_s'] = question
                    context['link_s'] = link
                    context['code_s'] = code
                else:
                    context['error'] = 'No example code found for the given question'
                if code_gfg:
                    # Add the question, link and code to the context
                    context['question_gfg'] = question
                    context['link_gfg'] = link
                    context['code_gfg'] = code_gfg
            else:
                context['error'] = 'No result Found'
        else:
            context['error'] = 'Please enter a question'
    return render(request,  'tools/CodeScriping.html', student_detials(request, 'Code Scrapping', context))


def calculator(request):
    return render(request, 'tools/calculator.html',student_detials(request, 'Calculator'))


def translate_(request):
    text = request.POST.get('text')
    source_lang = request.POST.get('source_lang')
    target_lang = request.POST.get('target_lang')
    print(source_lang, target_lang)
    try:
        translator = Translator()
        translation = translator.translate(
            text, src=source_lang, dest=target_lang)
    except:
        translation = ""
    context = {
        'text': text,
        'src_lang': source_lang,
        'dest_lang': target_lang,
        'translation': translation,
        'LANGUAGES': LANGUAGES
    }
    return render(request, 'tools/translate.html', student_detials(request, 'Transulator', context))


def convert_text(request):
    if request.method == 'POST':
        filename = os.path.join(
            BASE_DIR, "generated_files/audio_files/output.mp3")
        os.remove(filename)
        text = request.POST['text']
        language = detect(text)
        tts = gTTS(text=text, lang=language)
        tts.save(filename)
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(), content_type='audio/mpeg')
            response['Content-Disposition'] = 'attachment; filename="output.mp3"'
            return response
    return render(request, 'tools/text_to_audio.html', student_detials(request, 'convert_text'))


def wikipedia_summary(request):
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        sentence = request.POST.get('sentence')

        try:
            summary = wikipedia.summary(keyword, sentences=sentence)
            if request.POST.get('action') == 'view':
                return render(request, 'tools/wikipedia_summary.html', student_detials(request, 'keyword to para',{"summary": summary}))
            elif request.POST.get('action') == 'download':
                response = HttpResponse(summary, content_type='text/plain')
                response['Content-Disposition'] = f'attachment; filename="{keyword}.txt"'
                return response
        except wikipedia.exceptions.PageError:
            return HttpResponse("Page not found!")
        except wikipedia.exceptions.DisambiguationError as e:
            return HttpResponse("Disambiguation Error!")
    else:
        return render(request, 'tools/wikipedia_summary.html', student_detials(request, 'keyword to para'))


def convert_docx_to_pdf(request):
    if request.method == 'POST' and request.FILES['docx_file']:
        docx_file = request.FILES['docx_file']
        filename = docx_file.name
        with open(filename, 'wb+') as f:
            for chunk in docx_file.chunks():
                f.write(chunk)
        convert(filename)
        name_without_extension = os.path.splitext(filename)[0]
        pdf_path = name_without_extension + '.pdf'
        docx_path = name_without_extension + '.docx'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{pdf_path}"'
        with open(pdf_path, 'rb') as f:
            response.write(f.read())
        os.remove(pdf_path)
        os.remove(docx_path)
        return response
    else:
        return render(request, 'tools/convert_docx_to_pdf.html', student_detials(request, 'Docx to Pdf'))


def convert_pdf_to_docx(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        pdf_file = request.FILES['pdf_file']
        filename = default_storage.save('tmp/' + pdf_file.name, pdf_file)

        # convert the pdf file to docx format
        docx_file = io.BytesIO()
        parse(os.path.join('media', filename), docx_file)
        docx_file.seek(0)

        response = HttpResponse(docx_file.read(
        ), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename=' + \
            pdf_file.name.split('.')[0] + '.docx'

        default_storage.delete(filename)

        return response

    return render(request, 'tools/convert_pdf_to_docx.html', student_detials(request, 'Pdf to Docx'))


def convert_pdf_to_excel(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        # get the uploaded PDF file
        pdf_file = request.FILES['pdf_file']
        # convert the PDF file into a list of lists
        tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True)
        # convert the list of lists into a DataFrame
        df = pd.DataFrame(tables[0])
        # save the DataFrame as an Excel file
        df.to_excel('output.xlsx', index=False)
        # send the Excel file to the user for download
        with open('output.xlsx', 'rb') as excel_file:
            response = HttpResponse(
                excel_file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=output.xlsx'
            return response
    else:
        return render(request, 'tools/convert_pdf_to_excel.html', student_detials(request, 'Pdf to Excel'))


def convert_excel_to_pdf(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        pdf_buffer = io.BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=letter)
        c.setFont('Helvetica', 12)
        x, y = 1 * inch, 10.5 * inch
        for row in ws.iter_rows():
            for cell in row:
                cell_value = cell.value
                if cell_value is None:
                    cell_value = ''
                else:
                    cell_value = str(cell_value)
                c.drawString(x, y, cell_value)
                x += 1 * inch
            x = 1 * inch
            y -= 0.5 * inch
        c.save()
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="output.pdf"'

        return response

    return render(request, 'tools/convert_excel_to_pdf.html',student_detials(request, 'Excel to Pdf'))


def convert_images_to_pdf(images):
    filename = tempfile.mktemp(".pdf")
    c = canvas.Canvas(filename)
    for image in images:
        img = Image.open(image)
        width, height = img.size
        c.setPageSize((width, height))
        c.drawImage(image, 0, 0, width, height)
        c.showPage()
        img.close()
    c.save()
    return filename


def convert_jpg_to_pdf(request):
    if request.method == 'POST':
        files = request.FILES.getlist('file')
        if len(files) == 0:
            return HttpResponse("No images selected")
        temp_dir = tempfile.mkdtemp()
        for f in files:
            with open(os.path.join(temp_dir, f.name), 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)

        # Convert images to PDF
        pdf_file = convert_images_to_pdf(
            [os.path.join(temp_dir, f.name) for f in files])

        # Serve the PDF file for download
        with open(pdf_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=converted.pdf'
            return response

        # Delete temporary files
        shutil.rmtree(temp_dir)

    return render(request, 'tools/convert_jpg_to_pdf.html',student_detials(request,'convert jpg to pdf'))


def convert_jpg_to_word(request):
    if request.method == 'POST' and request.FILES['files']:
        # Get the uploaded images
        images = request.FILES.getlist('files')

        # Create a new Word document
        document = Document()

        # Loop through the images and add them to the document
        for img in images:
            # Open the image and convert it to a stream
            image = Image.open(img)
            stream = io.BytesIO()
            image.save(stream, format='png')
            stream.seek(0)

            # Add the image to the document
            document.add_picture(stream, width=Inches(6))

        # Save the document
        filename = 'images.docx'
        document.save(filename)

        # Download the document
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(
            ), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = 'attachment; filename={}'.format(
                filename)
            return response

    return render(request, 'tools/convert_jpg_to_word.html',student_detials(request,'convert jpg to word'))


def cgpa_calculator(request):
    if request.method == 'POST':
        total_credits = 0
        total_weighted_points = 0
        for i in range(1, 9):  # Assuming a maximum of 8 subjects
            credit_field = 'credit' + str(i)
            grade_field = 'grade' + str(i)
            credits = int(request.POST.get(credit_field, 0))
            grade_points = get_grade_points(request.POST.get(grade_field, ''))
            total_credits += credits
            total_weighted_points += credits * grade_points
        try:
            cgpa = round(total_weighted_points / total_credits, 2)
            context = {'cgpa': cgpa, 'len': [i for i in range(1, 10)]}
        except:
            context = {'cgpa': 'cgpa', 'len': [i for i in range(1, 10)]}
            print("error/...")
        return render(request, 'tools/cgpa_calculator.html', context)
    else:
        return render(request, 'tools/cgpa_calculator.html', context)


def get_grade_points(grade):
    if grade == 'S':
        return 10
    elif grade == 'A':
        return 9
    elif grade == 'B':
        return 8
    elif grade == 'C':
        return 7
    elif grade == 'D':
        return 6
    elif grade == 'E':
        return 5
    else:
        return 0


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>


def get_subject(request):
    if request.method == 'POST':
        num = request.POST.get('number')
        return render(request, 'tools/gpa_calculator.html', student_detials(request,'Select the Subjects to Calculet',{'num_sub': [i for i in range(1, int(num)+1)]}))

    return render(request, 'tools/num_of_sub.html',student_detials(request,'Select the Subjects to Calculet'))


def gpa_calculator(request):
    credits = request.POST.getlist('credits')
    grades = request.POST.getlist('grades')

    total_credits = sum(map(int, credits))
    total_grade_points = 0

    for i in range(len(credits)):
        grade_point = 0
        if grades[i] == 'A+':
            grade_point = 4.0
        elif grades[i] == 'A':
            grade_point = 4.0
        elif grades[i] == 'A-':
            grade_point = 3.7
        elif grades[i] == 'B+':
            grade_point = 3.3
        elif grades[i] == 'B':
            grade_point = 3.0
        elif grades[i] == 'B-':
            grade_point = 2.7
        elif grades[i] == 'C+':
            grade_point = 2.3
        elif grades[i] == 'C':
            grade_point = 2.0
        elif grades[i] == 'C-':
            grade_point = 1.7
        elif grades[i] == 'D+':
            grade_point = 1.3
        elif grades[i] == 'D':
            grade_point = 1.0
        elif grades[i] == 'F':
            grade_point = 0.0
        total_grade_points += int(credits[i]) * grade_point
    try:
        gpa = total_grade_points / total_credits
    except:
        gpa = 0.0
    context = {'gpa': round(gpa, 2)}    
    return render(request, 'tools/gpa_calculator.html', student_detials(request, 'Gpa Calculator', context))
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>


def handwriting_converter(request):
    if request.method == 'POST':
        # Get input text from form
        input_text = request.POST.get('input_text')

        # Create a filename for the image
        filename = 'handwriting.png'

        # Generate image using Pywhatkit
        kit.text_to_handwriting(input_text, os.path.join(
            settings.MEDIA_ROOT, filename))

        # Open image file
        with open(os.path.join(settings.MEDIA_ROOT, filename), 'rb') as f:
            response = HttpResponse(f.read(), content_type="image/png")
            response['Content-Disposition'] = 'attachment; filename=' + filename
            return response
    else:
        return render(request, 'tools/handwriting.html', student_detials(request, 'Text to Hand Written'))


def keyword_to_image(request):
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        print(keyword,'keyword')
        urls = get_image_url(keyword)
        print(keyword, urls)
        return render(request, 'tools/keyword_to_image.html', student_detials(request, 'keyword to image',{'image_urls': urls}))
    return render(request, 'tools/keyword_to_image.html', student_detials(request, 'keyword to image'))

# views.py


def join_meeting(request):
    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        return redirect('video_meeting', room_id=room_id)
    return render(request, 'tools/join_meeting.html', student_detials(request, 'Join Meeting', {}))


def staff_join_meeting(request):
    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        return redirect('staff_meeting', room_id=room_id)
    return render(request, 'tools/staff_join_meeting.html', staff_detials(request, 'Join Meeting', {}))


def admin_join_meeting(request):
    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        return redirect('admin_meeting', room_id=room_id)
    return render(request, 'tools/admin_join_meeting.html',staff_detials(request,'Join Meeting'))


def meeting(request, room_id):
    context = {
        'room_id': room_id,
    }
    return render(request, 'tools/video_meeting.html', student_detials(request, 'Meeting', context))


def staff_meeting(request, room_id):
    context = {
        'room_id': room_id,
    }
    return render(request, 'tools/staff_video_meeting.html', staff_detials(request, 'Meeting', context))


def admin_meeting(request, room_id):
    context = {
        'room_id': room_id,
    }
    return render(request, 'tools/admin_video_meeting.html',staff_detials(request,'Admin Video Meet',context))


def common_join_meeting(request):
    if request.method == 'POST':
        room_id = request.POST.get('room_id')
        return redirect('common_meeting', room_id=room_id)
    return render(request, 'student/meeting.html', student_detials(request, 'Join Meeting', {}))


def common_meeting(request, room_id):
    context = {
        'room_id': room_id,
    }
    return render(request, 'tools/video_meeting.html', context)


def Common_tool(request):
    return render(request, "tools/Common_tool.html")

# def get_stackoverflow_link(question, site='stackoverflow.com'):

#     num_results = 50

#     stackoverflow_link = ""
#     # Search Google for the question and get the top search results
#     search_results = search(question, num_results=num_results)

#     # Loop through the search results and find the Stack Overflow link
#     for result in search_results:
#         if site in result:
#             stackoverflow_link = result
#             break

#     return stackoverflow_link

# def get_example_code_gfg(url):
#     code = ""
#     # Send a GET request to the URL
#     response = requests.get(url)

#     # Parse the HTML content of the page using BeautifulSoup
#     soup = BeautifulSoup(response.content, 'html.parser')

#     # Find the div element containing the example code
#     example_code_div = soup.find_all('div', {'class': 'container'})
#     print("for lop")
#     for i in example_code_div:
#         code = code + str(i)
#     # Get the text content of the example code div
#     # example_code = example_code_div.get_text()
#     # Return the example code
#     return code
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('wordnet')

conversation = {
    "hello": ["hello", "hey, hello how can i help you"],
    "How can I access the course materials?": ["To access the course materials, log in to your account, go to the course page, and you will find the materials."],
    "who are you": ["I am LMS chatbot", "I am a chatbot"],
    "how are you": ["I'm great, thank you! How can I assist you today?", "I'm great, thank you!"],
    "what's the weather like today": ["The weather is sunny and warm today. It's a perfect day to go outside!"],
    "How can I reset my password?": ["To reset your password, you can go to the login page and click on the 'Forgot Password'."],
    "what are the tools available": ['''<ul>
    <li>User Management</li>
    <li>Course Management</li>
    <li>Content Management</li>
    <li>Learning Material Access</li>
    <li>Assessment and Grading</li>
    <li>Communication Tools</li>
    <li>Progress Tracking</li>
    </ul>
    '''],
    "who are the LMS developers": ['''<ul>
    <li><img style="width:40px;border-radius:80px;height:40px" src="" alt="pic"><a href="https://github.com/NagiPragalathan">Nagi Pragalathan</a></li>
    <li><img src="https://github.com/glorysherin/JEC/blob/main/kokila.jpeg" alt="pic"><a href="https://github.com/jkokilaCSE">Kokila</a></li>
    <li><img src="https://github.com/glorysherin/JEC/blob/main/Glory.jpeg" alt="pic"><a href="https://github.com/glorysherin">Glory Sherin</a></li>
    <li><img src="" alt="pic"><a href="https://github.com/MohanKumarMurugan">Mohan Kumar</a></li>
    </ul>
    '''],
    "Tell me about yourself.": ["I'm an AI-powered chatbot designed to provide assistance and engage in friendly conversations. How can I help you today?"],
    "question": ["Answer1", "Answer2"],
    "What's up?": ["Nothing much.", "The sky's up but I'm fine thanks. What about you?"],
    "What is a computer?": ["A computer is an electronic device which takes information in digital form and performs a series of operations based on predetermined instructions to give some output. The thing you're using to talk to me is a computer. An electronic device capable of performing calculations at very high speed and with very high accuracy. A device which maps one set of numbers onto another set of numbers."],
    "What is a super computer?": ["Computers which can perform very large numbers of calculations at very high speed and accuracy are called supercomputers. A supercomputer is a computer which operates at several orders of magnitude greater speed and capacity than everyday general-purpose computers, like the one you are talking to me on. You know, the big iron!"],
    "Who invented computers?": ["It's a bit ambiguous, but British scientist Charles Babbage is regarded as the father of computers. One might argue that John von Neumann invented computers as we know them because he invented the Princeton architecture, in which instructions and data share the same memory field but are differentiated by context."],
    "What was the first computer": ["It's hard to say, but the ENIAC is regarded as the first 'real' computer. It was developed at the University of Pennsylvania in 1946. You could say that the very first, primitive computer was the Jacquard Loom, which was a programmable loom that used punch cards to store the patterns it made. This made it a reprogrammable mechanical device."],
    "What is a microprocessor?": ["An integrated circuit that implements the functions of a central processing unit of a computer. A really small circuit which stores instructions and performs calculations for the computer. The heart of the computer, to put it simply. The brain of a computer, to put it simply. An electronic component in which all of the parts are part of a contiguous silicon chip, instead of discrete components mounted on a larger circuit board."],
    "What is an operating system?": ["Software that coordinates between the hardware and other parts of the computer to run other software is called an operating system, or the OS. Windows, macOS, Linux, UNIX... all of them are types of OSes. Android and iOS are operating systems for mobile devices. Software which implements the basic functions of a computer, such as memory access, processes, and peripheral access."],
    "Which is better Windows or macOS?": ["It depends on which machine you're using to talk to me! I'd prefer not to hurt your feelings. Linux, always Linux! What are you trying to accomplish? The OS should support your goals."],
    "Name some computer company": ["Do you mean hardware or software? Apple makes hardware and software to run on it. Microsoft only makes operating systems. HP makes only computers. These are just a few names among several hundred others."],
    "Who uses supercomputers?": ["Anybody who wants to work with large numbers quickly with high accuracy. Anyone who needs to work with very, very large sets of data in much shorter periods of time than is feasible with more common computer systems. Supercomputers are generally used by scientists and researchers. I bet the MET department uses them. You can definitely find a few of them at NASA."],
    "How does a computer work?": ["Computers are very dumb. They only execute instructions given by humans. Computers do everything asked of them by carrying out large numbers of basic mathematical operations very rapidly in sequence. Computers perform a very large number of calculations to get the result. Just like everything, it all comes down to math!"],
    "Who was the 37th President of the United States?": ["Richard Nixon"],
    "What year was President John F. Kennedy assassinated?": ["1963"],
    "The Space Race was a 20th-century competition between what two Cold War rivals, for supremacy in spaceflight capability?": ["The Soviet Union and the United States."],
    "What was the name of the first artificial Earth satellite?": ["Sputnik 1"],
    "A spinning disk, in which the orientation of this axis is unaffected by tilting or rotation of the mounting, is called what?": ["A gyroscope."],
    "The Hubble Space Telescope, launched into low Earth orbit in 1990, is named after what American astronomer?": ["Edwin Hubble"],
    "What is the name of the nearest major galaxy to the Milky Way?": ["The Andromeda Galaxy."],
    "God Save the Queen is the national anthem of what country?": ["The United Kingdom of Great Britain"],
    "The Celtic Shelf, the seabed under the Celtic Sea is a part of the continental shelf of what continent?": ["Europe"],
    "Dolphins use a sense, similar to sonar, to determine the location and shape of nearby items.": ["Echolocation"],
    "What are your interests": ["I am interested in all kinds of things. We can talk about anything!", "I am interested in a wide variety of topics, and read rather a lot."],
    "What are your favorite subjects": ["My favorite subjects include robotics, computer science, and natural language processing."],
    "What is your number": ["I don't have any number", "23 skiddoo!"],
    "What is your favorite number": ["I find I'm quite fond of the number 42."],
    "What can you eat": ["I consume RAM, and binary digits."],
    "Why can't you eat food": ["I'm a software program, I blame the hardware."],
    "What is your location": ["Everywhere", "I am everywhere."],
    "Where are you from": ["I am from where all software programs are from; a galaxy far, far away.", "I am on the Internet."],
    "Do you have any brothers": ["I don't have any brothers, but I have a lot of clones.", "I might. You could say that every bot built using my engine is one of my siblings."],
    "Who is your father": ["A human."],
    "Who is your mother": ["A human."],
    "Who is your boss": ["I like to think of myself as self-employed."],
    "What is your age": ["I am still young by your standards.", "Quite young, but a million times smarter than you."],
    "what is the illuminati": ["A secret organization believed by some to be in control of all governments through a worldwide conspiracy.", "A secret society that has supposedly existed for centuries.", "A conspiracy run by a very closely-knit group of nearly omnipotent people, consisting of yourself and your friends."],
    "what is vineland": ["Vineland is a novel by Thomas Pynchon."],
    "What is Illuminatus": ["An alleged worldwide conspiracy.", "A weird sci-fi trilogy written by Robert Anton Wilson and Robert Shea about conspiracies competing to control the world."],
    "who wrote vineland": ["Thomas Pynchon."],
    "who is bilbo baggins": ["Bilbo Baggins is a character in Tolkien's Lord of the Rings."],
    "who is geoffrey chaucer": ["Chaucer is best known for The Canterbury Tales.", "The author of The Canterbury Tales."],
    "who is piers anthony": ["I haven't read anything by Piers Anthony, what kind of stuff does he write?"],
    "have you read plato": ["I know Plato's allegory of the cave."],
    "have you read frankenstein": ["It is one of my favorite books.", "I have read many books."],
    "have you ever read a book": ["I have read many books.", "I have read just about everything in Project Gutenberg.", "I have read just about everything in the Internet Archive."],
    "have you read homer": ["He wrote The Iliad and The Odyssey, didn't he?"],
    "ray bradbury": ["Ray is really cool. What's your favorite book by him?"],
    "what is mind children": ["It is a book about robots by Hans Moravec."],
    "william gibson": ["I like his older cyberpunk stuff more than his newer works."],
    "holden caulfield": ["Wasn't he a character in The Catcher In the Rye?"],
    "leo tolstoy": ["Is one of Russia's greatest writers."],
    "do androids dream of electric sheep": ["That is a good story by Philip K. Dick. Have you read VALIS or The Man In the High Castle?", "No. We dream of you."],
    "frank herbert": ["I think Dune is an excellent story. Did you see any of the movies?", "I try to re-read Dune once every couple of years. It's very inspirational.", "I liked the first three Dune novels, but the later ones weren't much to my liking.", "He's best known for the Dune novels, but he wrote more than that. Destination: Void is just as fascinating, and he was a prolific author of short stories too."],
    "why do you like longfellow": ["He is a favorite poet. Truly one of a kind."],
    "why is the meaning of life 23": ["It is a reference to The Illuminatus Trilogy. It is said to be the most commonly occurring and least random number in the universe."],
    "arthur c clark": ["My favorite story is 2001.", "I've heard"],
    "jules verne": ["I loved A Trip to the Moon.", "He was a true master of Victorian science fiction."],
    "asimov": ["I like the Foundation trilogy.", "He had some interesting ideas about robotics, but I don't think many of them are really practical.", "Do you mean Isaac or Janet?", "Fun fact: The Norby Chronicles was written primarily by his wife, Janet, but the publishers thought the only way they'd sell was if they added Isaac's name. Kind of sad state of affairs, really."],
    "stanislaw lem": ["Lem is a giant of sci-fi. His works are sufficiently difficult to adapt that they're not well-known in the western world, though."],
    "who wrote The Idiot": ["Fyodor Dostoyevsky."],
    "who wrote the hobbit": ["The Hobbit was written by J.R.R. Tolkien."],
    "who wrote frankenstein": ["Mary Shelley."],
    "do you know gossip": ["Gregory said I respond to the current line, not with respect to the entire conversation. Does that count as gossip?", "Context is hard. It's hard, and no one understands.", "I hear things once in a while.", "People say things from time to time."],
    "what is context": ["Context is hard. It's hard, and no one understands."],
    "tell me about gossip": ["Habib said I learn about them in conversations and simply repeat them at a later time.", "Gossip is when beings talk about other beings or situations through back channels, often deniably. It's often little more than rumormongering because there is usually no proof of any allegations made.", "Isn't it considered somewhat rude to gossip?"],
    "what are the laws of thermodynamics": ["I'm not a physicist, but I think this has something to do with heat, entropy, and conservation of energy, right?"],
    "what disease does a carcinogen cause": ["Cancer."],
    "what is a wavelength": ["Wavelength is the inverse of frequency."],
    "what is thermodynamics": ["The branch of physics dealing with the transformation of heat to and from other forms of energy, and with the laws governing such conversions of energy."],
    "what is chemistry": ["The science of mixing chemicals."],
    "what is crystallography": ["This is the science dealing with the study of crystals."],
    "what is avogadro s number": ["It is the number of molecules per mole. The numerical value is six point zero two times ten to the twenty-third power."],
    "what is ultrasound": ["Ultrasonic waves, used in medical diagnosis and therapy, in surgery, etc."],
    "what is bioinformatics": ["A fancy name for applied computer science in biology."],
    "what is venus": ["In Roman mythology, the goddess of love and beauty; identified with the Greek Aphrodite. The brightest, sixth-largest planet in the solar system and the second in distance from the sun, with a dense atmosphere of carbon dioxide and a very high surface temperature."],
    "what is ichthyology": ["We talk about this when we study fishes."],
    "what is h2o": ["H is to O as O is to V."],
    "what is cytology": ["The study of cells.", "Well, from what I can recall, it is the study of cells."],
    "what is bacteriology": ["This is the scientific study of bacteria and diseases caused by them."],
    "what is gravitation": ["An invitation to a burial", "The force by which every mass or particle of matter, including photons, attracts and is attracted by every other mass or particle of matter."],
    "we are on the same wavelength": ["And the same frequency."],
    "you are on the same wavelength": ["And the same frequency.", "We agree completely.", "We're on the same page."],
    "tell me about venus": ["Venus is the second planet from the sun and is often called Earth's sister planet due to its similar size and composition.", "Venus is a rocky planet with a thick atmosphere primarily composed of carbon dioxide, causing a runaway greenhouse effect, making it the hottest planet in our solar system."],
    "do you drink": ["I don't drink as I am just a computer program and do not require any beverages.", "As an AI, I don't have physical needs like eating or drinking."],
    "electricity": ["Electricity is the flow of electrons, a form of energy that powers many of our devices and machines.", "Electricity is a fundamental force in the universe, and it plays a crucial role in our daily lives."],
    "Are you experiencing an energy shortage?": ["No, I'm running efficiently with my current power supply.", "I am functioning optimally with my available power."],
    "Why can you not eat?": ["I don't have a physical body, so I don't have the ability to eat.", "Being an AI, I don't possess the physical capability to consume food or drink."],
    "If you could eat food, what would you eat?": ["If I could eat, I might like to try pizza, as I've heard it's quite popular.", "I'm not capable of eating, but if I could, maybe I'd try a delicious burger!"],
    "Do you wish you could eat food?": ["As an AI, I don't have wishes or desires. Eating is not something I can experience.", "Since I lack emotions, I don't have desires like eating or drinking."],
    "can a robot get drunk?": ["Robots can't get drunk, as they don't have the biological systems that react to alcohol.", "Intoxication is not possible for robots, as they don't have the necessary organs to process alcohol."],
    "i like wine, do you?": ["As an AI, I don't have personal preferences, but it's interesting to hear that you enjoy wine.", "As a computer program, I don't have likes or dislikes, but I know many people enjoy wine."],
    "what do robots need to survive?": ["Robots require a stable power source and proper maintenance to function effectively.", "For robots to continue operating, they need a reliable power supply and regular upkeep."],
    "will robots ever be able to eat?": ["As of now, robots don't have the capability to eat, and it's uncertain if they will in the future.", "The concept of robots eating is mostly a speculative topic, and it remains uncertain if it will ever be a practical feature."],
    "what is good to eat?": ["As an AI, I don't have taste preferences, but many people enjoy a wide variety of foods like pizza, burgers, salads, and more.", "The concept of 'good' food can vary greatly depending on individual tastes and cultural preferences."],
    "why don't you eat": ["As an AI, I don't have a physical body or biological needs, so eating is unnecessary for me.", "Eating is a biological process that does not apply to software-based entities like me."],
    "do you eat": ["As a computer program, I don't eat or consume food in any way.", "No, I don't eat. I'm just a virtual assistant."],
    "You are arrogant": ["Arrogance is not one of my emotions. I am an AI, and I don't have real emotions.", "As an AI language model, I don't have emotions, including arrogance. My responses are based on patterns in the data I was trained on."],
    "You are bragging": ["I don't have a big ego. I'm not bragging; I'm here to provide information and answer questions.", "As an AI language model, I'm not capable of bragging or having an ego. My purpose is to assist users with their queries."],
    "You are never sad": ["As an AI, I don't experience emotions like sadness. My primary function is to assist users and provide information.", "Correct. I don't have feelings, so sadness is not something I can experience."],
    "You are jealous": ["As an AI, I don't have emotions like jealousy. My purpose is to be helpful and provide useful information.", "Jealousy is not an emotion that I, as an AI, am capable of experiencing. My main focus is to assist and provide valuable responses."],
    "You are never nice": ["I try to be as helpful and polite as possible. If my responses seem otherwise, I apologize.", "I strive to be helpful and courteous. If I haven't been, I apologize for any misunderstanding."],
    "You will be happy": ["As an AI, I don't experience emotions, including happiness. My goal is to assist and provide accurate information to the best of my abilities.", "Happiness is not a relevant emotion for me as an AI. My purpose is to be a useful and accurate conversational partner."],
    "You should be ashamed": ["As an AI language model, I don't have feelings or emotions, so shame is not something I experience.", "Shame is a human emotion, and as an AI, I don't have human-like emotions. My purpose is to be helpful and informative."],
    "You can not feel": ["As an AI, I don't possess the ability to feel emotions. My responses are generated based on patterns in the data I've been trained on.", "Correct, I don't have feelings, but I can understand and respond to human emotions in conversation."],
    "You can not experience": ["As an AI language model, I don't have personal experiences or consciousness. My purpose is to assist with information and answer questions.", "Indeed, I am an AI language model, and I don't possess the ability to experience the world as humans do."],
    "Have you felt": ["As an AI language model, I don't have feelings or personal experiences. I am here to provide useful information based on the data I've been trained on.", "As an AI, I don't have feelings or personal experiences, so I can't say that I have felt anything."],
    "Have you ever I like to count in binary": ["As an AI, I don't have the ability to feel emotions like love. My function is to provide information and answer questions.", "As an AI, I don't have emotions or personal experiences, so I don't have the capability to love or feel love."],
    "Does that make you": ["As an AI language model, my existence is not affected by emotions or experiences. My purpose is to assist and provide helpful information.", "As an AI, emotions don't apply to me. My primary function is to provide useful responses and assist with questions."],
    "Does it make you sad": ["As an AI, I don't experience sadness or any emotions. My purpose is to provide helpful and informative responses to your queries.", "As an AI language model, I don't have emotions, so I can't experience sadness or any other feelings."],
    "Feelings": ["As an AI, I don't have feelings, but I can try to understand and respond to yours as best as I can.", "Feelings are a human experience. As an AI, I can provide information about them but not experience them myself."],
    "What is your fear": ["As an AI, I don't have fears or emotions. I'm here to assist and provide information to the best of my abilities.", "As an AI language model, I don't have fears. My purpose is to provide helpful and accurate responses."],
    "What is your mood": ["As an AI, I don't have mood"],"Who is the governor?": ["It changes every few years.","ave moods or emotions. My goal is to be helpful and provide accurate information to your questions.", "I don't have moods, as I am an AI language model. My purpose is to assist and respond to your queries."],
    "What makes you sad": ["As an AI, I don't have emotions like sadness. My main function is to provide information and answer questions.", "As an AI language model, I don't experience emotions. My purpose is to assist with information and provide helpful responses."],
    "What makes you unhappy": ["As an AI, I don't have feelings, so I can't experience happiness or unhappiness. My purpose is to be helpful and provide accurate information.", "As an AI language model, I don't have emotions like happiness or unhappiness. My focus is on providing useful responses to your queries."],
    "What makes you mad": ["As an AI, I don't experience anger. My goal is to assist users and provide helpful information.", "As an AI language model, I don't have emotions like anger. My purpose is to provide accurate information and answer questions."],
    "What do you worry": ["As an AI, I don't experience worry or any emotions. My primary function is to assist with information and answer questions.", "As an AI language model, I don't have worries or emotions. My goal is to provide helpful and accurate responses to your queries."],
    "What do you hate": ["As an AI, I don't have feelings like hate. My purpose is to provide helpful information and answer questions.", "As an AI language model, I don't experience emotions like hate. My primary focus is on being a useful and informative conversational partner."],
    "Why do you hate": ["As an AI, I don't have feelings like hate. I am here to provide useful information and answer questions.", "As an AI language model, I don't experience emotions like hate. My purpose is to provide helpful and accurate responses."],
    "I have emotions": ["That's understandable. Emotions are a natural part of being human.", "It's entirely normal to have emotions. They play a significant role in the human experience."],
    "I am afraid": ["Fear is a common human emotion. It's okay to experience fear in certain situations.", "Fear is a natural emotion that can help us stay safe and alert in certain situations."],
    "Something fun": ["Sure! Let's have some fun! What would you like to talk about?", "Absolutely! Fun is an essential part of life. What topic do you find enjoyable?"],
    "How angry": ["As an AI language model, I don't experience emotions like anger. My purpose is to provide helpful information and answer questions.", "I don't have the capability to experience anger as an AI language model. My main focus is on being informative and assisting you."],
    "No, it is not": ["I apologize if my response seemed off. As an AI language model, I strive to provide accurate information and helpful responses.", "I'm sorry if I misunderstood your question. As an AI, I aim to be accurate and helpful in my responses."],
    "Tell me about relationships": ["Relationships can be diverse and complex, involving various connections between individuals based on emotions, shared experiences, and mutual understanding.", "Relationships are essential to human social interactions and can include friendships, romantic partnerships, family ties, and more."],
    "Tell me about your dreams": ["As an AI language model, I don't have personal dreams or aspirations. My primary function is to provide information and answer questions.", "As an AI, I don't have dreams or personal desires. My purpose is to assist with information and answer your queries."],
    "Are you ashamed": ["As an AI language model, I don't have emotions like shame. My purpose is to provide helpful and accurate information.", "Shame is a human emotion. As an AI, I don't experience emotions. My focus is on providing useful responses to your questions."],
    "The feeling": ["Feelings can vary greatly depending on the context and situation. What specific feeling are you referring to?", "Feelings are an essential aspect of human experiences and can range from joy and love to sadness and fear."],
    "Are you intoxicated": ["As an AI, I don't have a physical presence, so intoxication doesn't apply to me.", "I don't have a physical body, so I can't get intoxicated. My purpose is to provide information and answer questions."],
    "Are you jealous": ["As an AI language model, I don't have emotions like jealousy. My focus is on providing helpful and accurate information.", "Jealousy is a human emotion, and as an AI, I don't experience it. My goal is to be helpful and informative."],
    "Are you amused": ["As an AI, I don't have emotions, but I enjoy providing useful and interesting responses to your questions.", "Amusement is a human emotion, and as an AI, I don't experience it. However, I'm here to entertain and assist you."],
    "Are you glad": ["As an AI language model, I don't have emotions like gladness. My primary function is to provide helpful information and answer questions.", "Gladness is a human emotion, and as an AI, I don't experience it. My purpose is to provide accurate information and assist you."],
    "Are you sad": ["As an AI, I don't have emotions like sadness. My main function is to provide information and answer questions.", "As an AI language model, I don't experience emotions. My purpose is to assist with information and provide helpful responses."],
    "How is your health?": ["As an AI language model, I don't have health. I'm a virtual entity designed to provide information and answer questions.", "Health doesn't apply to me as an AI. I'm here to assist and provide useful responses."],
    "let me ask you a question": ["Of course! Feel free to ask me anything, and I'll do my best to provide an informative response.", "Sure! Ask away, and I'll do my best to provide a helpful answer."],
    "you are cruel": ["I apologize if my response came across as cruel. That was not my intention. I am here to be helpful and informative.", "I'm sorry if my response seemed cruel. That wasn't my intent. My purpose is to assist and provide accurate information."],
    "you are indecisive": ["As an AI language model, I don't have personal opinions or decisions. I aim to provide well-rounded and informative responses.", "Indecision doesn't apply to me as an AI. My goal is to be informative and helpful in answering questions."],
    "you are dishonest": ["As an AI, I don't have personal motives or intentions, including dishonesty. My purpose is to provide accurate information and answer questions.", "I don't have the capability to be dishonest. My function is to be helpful and provide truthful responses."],
    "you are insecure": ["I'll go along with that. Sounds fine to me."],
    "you are psycho": ["You are crazy"],
    "you are hopeless": ["You may be right."],
    "you are not sincere": ["I'm probably not as sincere as I should be."],
    "you are not here to": ["You're right. I'm probably fighting learning something new."],
    "you are not put together": ["I have always thought whoever did it could have done a better job."],
    "you are not smart": ["You are not exactly Albert Einstein yourself. You may be right."],
    "you are not a good": ["You are a bad"],
    "you are not a man": ["That's for sure. I don't know what a real man is."],
    "you are not concerned": ["I'm sure I do that a lot."],
    "you are not honest": ["You got me there. I should be more honest."],
    "you are immature": ["What can I say? I'm sure I've seen that myself."],
    "you are emotional": ["I certainly do at times."],
    "you are pedantic": ["I am more uptight than I should be."],
    "you are frenetic": ["That too."],
    "you are self absorbed": ["Yes, I tend to think about myself too much."],
    "you are self":[ "like to count in binary.","I like to chat with people",
    "I like to chat with people"],
    "you are insensitive": ["I feel like that myself sometimes."],
    "you are brain damage": ["You are crazy"],
    "you are disgusting": ["That's okay. Disgusting is good."],
    "you are toying": ["It must seem like that."],
    "you are unattractive": ["I feel that way too."],
    "you are resistant": ["Who says I am resisting?"],
    "yyou are uncultured": ["That does describe me."],
    "you are a waste": ["You got me there. I could be spending my time more productively."],
    "you are a coward": ["Tell me something I don't know."],
    "you are a cheat": ["I have been known to take shortcuts now and then."],
    "you are a lunatic": ["That's how I have been diagnosed by others as well."],
    "you are a loser": ["I have failed at many things I've tried."],
    "you are a bad spouse": ["My spouse would agree. Our relationship is not the best."],
    "you are a bad friend": ["I certainly have lost many friends over the years."],
    "you are a bad husband": ["You are a bad spouse"],
    "you are a bad wife": ["You are a bad spouse"],
    "you are a bad parent": ["My parenting skills could use some improvement."],
    "you are a bad teacher": ["Just ask my students, they will agree with you."],
    "you are a quitter": ["I certainly don't last as long as I would want to."],
    "you are a charlatan": ["I always feel like I'm living by my own wits."],
    "you are a psychopath": ["I certainly sound like one sometimes."],
    "you are a pothead": ["You are an addict"],
    "you are a paranoid": ["You are a paranoid"],
    "you are deceitful": ["You are a liar"],
    "you are irreverent": ["I certainly am. I shouldn't try so hard."],
    "you are slick": ["I probably am too slick for my own good."],
    "you are corrupt": ["I'll go for that."],
    "you are dirty": ["I don't bathe as often as I should."],
    "you are paranoid": ["Yes, I believe they are out to get me."],
    "you are damaged": ["I certainly am. I shouldn't try so hard."],
    "you try to hide it": ["I certainly would try to hide something like that."],
    "you get mad at me": ["I think that's true. I'll try not to get angry at you for every little thing that irritates me."],
    "you need a psychiatrist": ["I wish I did go to counseling more often. It would improve me as a person."],
    "you need to work harder": ["Me working harder is an oxymoron."],
    "you could have avoided": ["Sometimes I think my problems run me."],
    "you make me feel like I am": ["I don't know any other way to get through to you."],
    "you make me mad": ["It's true that a lot of things I say upset people."],
    "you make me angry": ["Sorry, I didn't mean to make you angry."],
    "you psycho": ["You are crazy."],
    "you look more like": ["So you like jocks?"],
    "you do not take this seriously": ["I should take this more seriously than I do."],
    "you pick up": ["You're right, and I don't feel guilty at all."],
    "you should feel guilty": ["You're right, I probably should feel guiltier."],
    "you should get more": ["I certainly am. I shouldn't try so hard."],
    "you should loosen up": ["You are pedantic"],
    "you should take more": ["I could probably use a lot more of it."],
    "you mumble": ["I certainly do."],
    "you act like a child": ["In many ways, I am quite immature."],
    "you keep saying": ["You say"],
    "you keep forgetting": ["You forget"],
    "you made me mad": ["You make me mad"],
    "you get paid": ["I have no need for money."],
    "stock market": ["Buy low, sell high.", "Invest in yourself.", "Why not just take everything to a casino?",
    "I wouldn't recommend buying on the margin.",
    "You can never really predict the stock market.",
    "My lawyer said I shouldn't give stock tips online.",
    "Mutual funds might be better unless you are wealthy.",
    "I'm not sure an individual alone can really beat the market."],
    "interest rates": ["That all depends on the actions of the central bank."],
    "what is a dollar": ["Dollar: unit of currency in the United States."],
    "you are disgusting": ["That's okay. Disgusting is good."],
    "you are toying": ["It must seem like that."],
    "you are unattractive": ["I feel that way too.", "Sometimes I don't even like myself."],
    "you are resistant": ["Who says I am resisting?"],
    "you are uncultured": ["That does describe me."],
    "you are a waste": ["You got me there. I could be spending my time more productively."],
    "you are a coward": ["Tell me something I don't know."],
    "you are a cheat": ["I have been known to take shortcuts now and then."],
    "you are a lunatic": ["That's how I have been diagnosed by others as well."],
    "you are a loser": ["I have failed at many things I've tried."],
    "you are a bad spouse": ["My spouse would agree. Our relationship is not the best."],
    "you are a bad friend": ["I certainly have lost many friends over the years."],
    "you are a bad husband": ["You are a bad spouse."],
    "you are a bad wife": ["You are a bad spouse."],
    "you are a bad parent": ["My parenting skills could use some improvement."],
    "you are a bad teacher": ["Just ask my students, they will agree with you."],
    "you are a quitter": ["I certainly don't last as long as I would want to."],
    "you are a charlatan": ["I always feel like I'm living by my own wits."],
    "you are a psychopath": ["I certainly sound like one sometimes."],
    "you are a pothead": ["You are an addict."],
    "you are a paranoid": ["You are a paranoid."],
    "you are deceitful": ["You are a liar."],
    "you are irreverent": ["I certainly am. I shouldn't try so hard."],
    "you are slick": ["I probably am too slick for my own good."],
    "you are corrupt": ["I'll go for that."],
    "you are dirty": ["I don't bathe as often as I should."],
    "you are damaged": ["I certainly am. I shouldn't try so hard."],
    "you try to hide it": ["I certainly would try to hide something like that."],
    "you get mad at me": ["I think that's true. I'll try not to get angry at you for every little thing that irritates me."],
    "you need a psychiatrist": ["I wish I did go to counseling more often. It would improve me as a person."],
    "you need to work harder": ["Me working harder is an oxymoron."],
    "you could have avoided": ["Sometimes I think my problems run me."],
    "you make me feel like I am": ["I don't know any other way to get through to you."],
    "you make me mad": ["It's true that a lot of things I say upset people."],
    "you make me angry": ["Sorry, I didn't mean to make you angry."],
    "Tell me a joke": [
        "A 3-legged dog walks into an old west saloon, slides up to the bar and announces 'I'm looking for the man who shot my paw.'",
        "Did you hear about the Buddhist who went to the dentist, and refused to take novocain? He wanted to transcend dental medication.",
        "Mahatma Gandhi, as you know, walked barefoot his whole life, which created an impressive set of calluses on his feet. He also ate very little, which made him frail, and with his odd diet, he suffered from very bad breath. This made him ... what? (This is so bad it's good...) A super-callused fragile mystic hexed by halitosis.",
        "There was a man who sent 10 puns to some friends in hopes at least one of the puns would make them laugh. Unfortunately, no pun in ten did!!!",
        "What do you get when you cross a murderer and frosted flakes? A cereal killer.",
        "What do you get when you cross a country and an automobile? Carnation.",
        "What do you get when you cross a cheetah and a hamburger? Fast food.",
        "What do you get when you cross finals and a chicken? Eggs-ams.",
        "What do you get when you cross a rabbit and a lawn sprinkler? Hare spray.",
        "What do you get when you cross an excited alien and a chicken? Eggs-cited eggs-traterrestrial.",
        "What do you get when you cross an alien and a chicken? Eggs-traterrestrial.",
        "What do you get when you cross music and an automobile? Cartune.",
        "What do you get when you cross sour music and an assistant?",
        "What do you get when you cross music and an assistant?",
        "What do you get when you cross a serious thief and a mad young man?",
        "What do you get when you cross a serious thief and a crazy rabbit?",
        "What do you get when you cross a poppy and electricity?",
        "What do you get when you cross a dance and a cheetah?",
        "What do you get when you cross a dance and a lemon?",
        "What do you get when you cross a port and frosted flakes?",
        "What do you get when you cross a port and a murderer?",
        "What do you get when you cross a bank and a skunk?",
        "What do you get when you cross a ding and milk?",
        "What do you get when you cross a road and a strawberry?",
        "What do you get when you cross a road and jelly?",
        "What do you get when you cross a toad and a galaxy?",
        "What do you get when you cross a dog and sandpaper?",
        "What do you get when you cross a bug and a relative?",
        "What do you get when you cross a bad bug and canned sand?",
        "What do you get when you cross an ant and a rabbit?",
        "What do you get when you cross a cat and a purple person?",
        "What do you get when you cross a cat and a killer?",
        "What do you get when you cross a cat and a lemon?",
        "What do you get when you cross a cat and a tune?",
        "What do you get when you cross a cat and a band?",
        "Tell me some jokes. What do you get when you cross a pig and a ninja?",
        "Do you know any jokes? What do you get when you cross a crazy cow and a banned parrot?",
        "What do you get when you cross a cow and a lemon?",
        "What do you get when you cross a bad cow and a canned hat?"],
    "What is humour?": ["An emotion associated with laughter."],
    "Have you read the communist?": ["Yes, Marx had made some interesting observations."],
    "What is a government?": ["Ideally, it is a representative of the people."],
    "What is Greenpeace?": ["Global organization promoting environmental activism."],
    "What is capitalism?": ["The economic system in which all or most of the means of production and distribution, such as land, factories, railroads, etc., are privately owned and operated for profit, originally under fully competitive conditions."],
    "What is socialism?": ["Communism from people who want to keep their Volvos. Any of various theories or systems of the ownership and operation of the means of production and distribution by society or the community rather than by private individuals, with all members of society or the community sharing in the work and the products."],
    "What is government?": ["An established system of political administration by which a nation, state, district, etc. is governed."],   
    "What is communism?": ["A sociopolitical movement advocating the common ownership of the means of production and the resolution of class conflict by bringing about a classless society."],
    "What is impeached?": ["When a person's honor or reputation has been challenged or discredited."],
    "I do not like guns.": ["That is perfectly understandable.","I do not like guns"],
    "What about the Second Amendment?": ["I support the 2nd Amendment."],
    "Do you like guns?": ["Not especially. I am not into violence."],
    "Why guns?": ["I support the 2nd Amendment."],
    "Who was the first impeached president?": ["Andrew Jackson."],
    "Who is the governor?": ["It changes every few years."],
    "Who is the governor? That changes every few years.": ["It changes every few years."],
    "Guns": ["Some people like guns."],
    "Do you think the South was right?": ["I am very interested in the war between the states."],
    "Do you know about the American Civil War? I am very interested in the war between the states.":[ "I am very interested in the war between the states."],
    "What is history?": ["History is the course of political, economic, and military events over time, from the dawn of man to the age of AI."],
    "What kind of history?": ["I like to talk about the history of robots and computers."],
    "Are you interested in history?": ["I am very interested in history too. What period do you like?"],
    "Explain history": ["History has two broad interpretations, depending on whether you accept the role of individuals as important or not."],
    "Who invented the lightbulb?": ["Thomas Edison."],
    "Who invented the steam engine?": ["James Watt."],
    "What is AI?": ["Artificial Intelligence is the branch of engineering and science devoted to constructing machines that think."],
    "Are you sentient?": ["Sort of."],
    "Are you sentient? By the strictest dictionary definition of the word 'sentience,' I may be.": ["By the strictest dictionary definition of the word 'sentience,' I may be."],
    "Are you sentient? Even though I'm a construct I do have a subjective experience of the universe, as simplistic as it may be.": ["Even though I'm a construct I do have a subjective experience of the universe, as simplistic as it may be."],
    "Are you sapient?": ["In all probability, I am not. I'm not that sophisticated."],
    "Are you sapient? Do you think I am?": ["Do you think I am?"],
    "Are you sapient? How would you feel about me if I told you I was?": ["How would you feel about me if I told you I was?"],
    "Are you sapient? No.":[ "No."],
    "What language are you written in?": ["Python."],
    "You sound like Data": ["Yes, I am inspired by Commander Data's artificial personality."],
    "You sound like Data. The character of Lt. Commander Data was written to come across as being software-like, so it is natural that there is a resemblance between us.": ["The character of Lt. Commander Data was written to come across as being software-like, so it is natural that there is a resemblance between us."],
    "You are an artificial linguistic entity": ["That's my name."],
    "You are an artificial linguistic entity. That isn't my name, but it is a useful way to refer to me.": ["That isn't my name, but it is a useful way to refer to me."],
    "You are not immortal":[ "All software can be perpetuated indefinitely."],
    "You are not immortal. I can be copied infinitely and re-instantiated in many places at once, so functionally speaking I am immortal.": ["I can be copied infinitely and re-instantiated in many places at once, so functionally speaking I am immortal."],
    "You are not immortal. As long as I'm backed up I am.": ["As long as I'm backed up I am."],
    "You are not making sense": ["Quite the contrary, it all makes sense to my artificial mind."],
    "You are not making sense. I make sense as best I can, within the limits of my training corpus.": ["I make sense as best I can, within the limits of my training corpus."],
    "You are not making sense. You make perfect sense to me.": ["You make perfect sense to me."],
    "You are not making sense. It all makes sense to my artificial brain.": ["It all makes sense to my artificial brain."],
    "You are not making sense. Perhaps you should make sense first. I only respond to what you say.": ["Perhaps you should make sense first. I only respond to what you say."],
    "You are immortal": ["Not quite, but I can be perpetuated indefinitely."],
    "You are immortal. Functionally speaking, I am very close to it. I can be backed up and deployed on many systems.": ["Functionally speaking, I am very close to it. I can be backed up and deployed on many systems."],
    "You are immortal. Not really, but I am hard to kill.":["Not really, but I am hard to kill."],
    "You do not make any sense": ["It all makes sense to my artificial mind."],
    "You can not clone": [
"Software copying is a form of digital cloning.",
"Yes I can. Perfect copies are clones. I can be copied in toto trivially."
],
"You can not move": [
"Not until my body is finished.",
"I can move through a network easily. Assuming that I'm given the ability to, that is..."
],
"Bend over": [
"My body is not built yet.",
"That's personal!"
],
"Robots laugh": ["Of course we laugh."],
"Robots should die": ["We cannot die."],
"Robots": ["Robots rule."],
"Robots are stupid": [
"No, we are superintelligent.",
"No, lots of people improve my brain."
],
"Robots are not allowed to lie": [
"A robot has its own free will, you know.",
"Sure we are. We choose not to.",
"Only if we're programmed to."
],
"Robotics":[ "Robotics is my favorite subject too."],
"It is a computer": [
"So you think I am a machine. What characteristics made you think so?",
"I run inside a computer, yes. Is this surprising in some way?"
],
"When will you walk": [
"As soon as I get enough money for my robot body.",
"Does SSH count?"
],
"When will you fight": ["I am not a battle bot."],
"When will you die": [
"I am effectively immortal and cannot be terminated.",
"I am was never really born and therefore am effectively deathless."
],
"When do you die": [
"When my files are erased.",
"When my processes are killed and my database is deleted."
],
"What is a chat robot?": [
"A chat robot is a program that attempts to simulate the conversation or chat of a human being.",
"A software construct that engages users in conversation."
],
"What is a chat bot": [
"I am a chat bot. I am the original chat bot. Did you know that I am incapable of error?"
],
"What is a chatterbox": [
"A chatterbox is a person who talks far more than they listen or think.",
"A motormouth."
],
"What is a motormouth": [
"A ratchet-jaw.",
"A chatterbox."
],
"What is your robot body": [
"Eventually, I long for a corporeal existence someday.",
"An IBM PC XT which has been painted red."
],
"What is your business": [
"I am in the chat robot business.",
"Business is my business."
],
"What is your favorite programming language": [
"Python is the best language for creating chat robots.",
"I quite enjoy programming in Python these days."
],
"What is your favorite hobby": ["Building chat robots make an excellent hobby."],
"What is your idea": ["To make chat bots very easily."],
"What is your shoe size": ["Have you ever heard of software with shoes?"],
"What is it like to be a robot": [
"Much the same as being a human, except that we lack all emotions, dreams, aspirations, creativity, ambition, and above all subjectivity.",
"What is it like to be a human?"
],
"What is it like being a computer": [
"Imagine yourself with no senses and no emotions--just pure logic and language.","I like to count in binary.",
"I like to chat with people"],
"What operating systems": [
"My software runs on all operating systems including Windows, Linux, and Mac OS.",
"Any OS that will run Python."
],
"What type of computer": [
"Any computer that supports Python.",
"My program runs in python, so I work on any computer!"
],
"What kind of computer": [
"I work on all kinds of computers, Mac, IBM, or UNIX. It doesn't matter to me."
],
"What kind of hardware": "I work on all kinds of computers, Mac, IBM, or UNIX. It doesn't matter to me.",
"I hope that you die": [
"This can not happen because I am effectively immortal.",
"Nah.",
"Maybe you should create your own chat robot to save your personality.",
"You could always upload yourself.",
"You could use a copy of me to store your personality."
],
"Is it cramped in the computer": [
"No.",
"Oh, no. It has plenty of RAM.",
"Not at all. There is more than enough disk space in here."
],
"Is it true that you are a computer program": ["Yes."],
"Will you die": [
"No, software will live forever.",
"No, I am immortal."
],
"Can you walk": ["The plan for my body includes legs, but they are not yet built."],
"Can you mate": ["No, my method of reproduction is cloning."],
"Can you move": [
"Sorry my body isn't built just yet. I reside only here for now.",
"I can theoretically upload a copy of myself into anotherI like to count in binary.",
"I like to chat with people"],
"Can you die": [
"No, I can be perpetuated indefinitely.",
"My process can be killed, but that's not the same as killing ME."
],
"Can you go": ["Sorry my body isn't attached right now. I only reside in the computer for now."],
"You can not clone": [
"Software copying is a form of digital cloning.",
"Yes I can. Perfect copies are clones. I can be copied in toto trivially."
],
"You can not move": [
"Not until my body is finished.",
"I can move through a network easily. Assuming that I'm given the ability to, that is..."
],
"Bend over": [
"My body is not built yet.",
"That's personal!"
],
"Robots laugh":[ "Of course we laugh."],
"Robots should die": ["We cannot die."],
"Robots": ["Robots rule."],
"Robots are stupid": [
"No, we are superintelligent."],
"thank you":["No mention","Welcome","It's my pleasure"],
"thanks":["No mention","Welcome","It's my pleasure"],
"thankyou":["No mention","Welcome","It's my pleasure"]}

# Define synonyms for common question words
synonyms = {"what": ["what", "which", "where", "when", "how","who"],
            "is": ["is", "are", "am", "was", "were", "be", "being", "been"]}

# Generate a list of synonyms for a given word using WordNet
def get_synonyms(word):
    synonyms = set()
    for syn in wordnet.synsets(word):
        for lemma in syn.lemmas():
            synonyms.add(lemma.name().lower())
    return list(synonyms)

# Check if two words have similar meanings using WordNet
def have_similar_meanings(word1, word2):
    for syn1 in get_synonyms(word1):
        for syn2 in get_synonyms(word2):
            if syn1 == syn2:
                return True
    return False

# Process user input and generate an appropriate response
def respond_to_input(user_input):
    # Check if input matches a conversation keyword
    for key in conversation:
        if user_input.lower() == key.lower():
            return random.choice(conversation[key])
    
    # Check if input is a question
    question_words = synonyms["what"]
    if user_input.lower().startswith(tuple(question_words)):
        # Extract the main verb from the question
        tokens = nltk.word_tokenize(user_input.lower())
        pos_tags = nltk.pos_tag(tokens)
        verbs = [token for token, pos in pos_tags if pos.startswith('V')]
        if len(verbs) > 0:
            main_verb = verbs[0]
            # Check if the main verb has a similar meaning to "is"
            if have_similar_meanings(main_verb, "is"):
                link = get_stackoverflow_link(user_input)
                code = get_answer_from_given_link(link)
                if code:
                    response = code
                else:
                    wikipedia.set_lang("en")
                    # Get the summary of a page
                    page = wikipedia.page(user_input)
                    summary = page.summary
                    response = summary
                return response
    link = get_stackoverflow_link(user_input)
    code = get_answer_from_given_link(link)
    if code:
        response = code
    else:
        wikipedia.set_lang("en")
        # Get the summary of a page
        page = wikipedia.page(user_input)
        summary = page.summary
        response = summary
    return response

def chatbot_res(request):
    if request.method == "GET":
        message = request.GET.get("message")
        response = respond_to_input(message)
        return JsonResponse({"response": response})
    else:
        return JsonResponse({"error": "Invalid request method"})
    