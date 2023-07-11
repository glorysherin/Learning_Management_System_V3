import pywhatkit as kit
from django.conf import settings
from django.shortcuts import render
from googletrans import Translator, LANGUAGES
from django.http import HttpResponse
from gtts import gTTS
from langdetect import detect
import os
import io
from LMS.settings import BASE_DIR
import wikipedia
from docx2pdf import convert
from django.core.files.storage import default_storage
from pdf2docx import parse
import tabula
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from PIL import Image
import tempfile
from pdf2docx import Converter
from docx import Document
from docx.shared import Inches
from .Tool.Code_scriping_Tool import get_image_url

from django.shortcuts import render
from .Tool.Code_scriping_Tool import get_stackoverflow_link,get_stackoverflow_link_1, get_example_code_gfg, get_answer_from_given_link


def Common_toolHome(request):
    return render(request, "Common_Page_Tools/Common_tool.html")


def Common_base(request):
    return render(request, "Common_Page_Tools/Common_base.html")


def Common_Code_scriping(request):
    context = {}
    if request.method == 'POST':
        question = request.POST.get('question')
        if question:
            # Get the Stack Overflow link for the question
            link = get_stackoverflow_link_1(question)
            link_gfg = get_stackoverflow_link_1(question, 'geeksforgeeks.org')
            if link:
                # Get the example code from the link
                code = get_answer_from_given_link(link)
                code_gfg = get_example_code_gfg(link_gfg)
                if code:
                    # Add the question, link and code to the context
                    context['question_s'] = question
                    context['link_s'] = link
                    context['code_s'] = code
                else:
                    context['error'] = 'No example code found for the given question'
                if code_gfg:
                    # Add the question, link and code to the context
                    context['question_gfg'] = question
                    context['link_gfg'] = link
                    context['code_gfg'] = code_gfg
            else:
                context['error'] = 'No Stack Overflow link found for the given question'
        else:
            context['error'] = 'Please enter a question'
    return render(request, 'Common_Page_Tools/CodeScriping.html', context)


def Common_calculator(request):
    return render(request, 'Common_Page_Tools/calculator.html')


def Common_translate_(request):
    text = request.POST.get('text')
    source_lang = request.POST.get('source_lang')
    target_lang = request.POST.get('target_lang')
    print(source_lang, target_lang)
    try:
        translator = Translator()
        translation = translator.translate(
            text, src=source_lang, dest=target_lang)
    except:
        translation = ""
    context = {
        'text': text,
        'src_lang': source_lang,
        'dest_lang': target_lang,
        'translation': translation,
        'LANGUAGES': LANGUAGES
    }
    return render(request, 'Common_Page_Tools/translate.html', context)


def Common_convert_text(request):
    if request.method == 'POST':
        filename = os.path.join(
            BASE_DIR, "generated_files/audio_files/output.mp3")
        try:
            os.remove(filename)
        except:
            pass
        text = request.POST['text']
        language = detect(text)
        tts = gTTS(text=text, lang=language)
        tts.save(filename)
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(), content_type='audio/mpeg')
            response['Content-Disposition'] = 'attachment; filename="output.mp3"'
            return response
    return render(request, 'Common_Page_Tools/text_to_audio.html')


def Common_wikipedia_summary(request):
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        sentence = request.POST.get('sentence')

        try:
            summary = wikipedia.summary(keyword, sentences=sentence)
            if request.POST.get('action') == 'view':
                return render(request, 'Common_Page_Tools/wikipedia_summary.html', {"summary": summary})
            elif request.POST.get('action') == 'download':
                response = HttpResponse(summary, content_type='text/plain')
                response['Content-Disposition'] = f'attachment; filename="{keyword}.txt"'
                return response
        except wikipedia.exceptions.PageError:
            return HttpResponse("Page not found!")
        except wikipedia.exceptions.DisambiguationError as e:
            return HttpResponse("Disambiguation Error!")
    else:
        return render(request, 'Common_Page_Tools/wikipedia_summary.html')


def Common_convert_docx_to_pdf(request):
    if request.method == 'POST' and request.FILES['docx_file']:
        docx_file = request.FILES['docx_file']
        filename = docx_file.name
        with open(filename, 'wb+') as f:
            for chunk in docx_file.chunks():
                f.write(chunk)
        convert(filename)
        name_without_extension = os.path.splitext(filename)[0]
        pdf_path = name_without_extension + '.pdf'
        docx_path = name_without_extension + '.docx'
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{pdf_path}"'
        with open(pdf_path, 'rb') as f:
            response.write(f.read())
        os.remove(pdf_path)
        os.remove(docx_path)
        return response
    else:
        return render(request, 'Common_Page_Tools/convert_docx_to_pdf.html')


def Common_convert_pdf_to_docx(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        pdf_file = request.FILES['pdf_file']
        filename = default_storage.save('tmp/' + pdf_file.name, pdf_file)

        # Convert the PDF file to DOCX format
        docx_file = io.BytesIO()
        pdf_path = default_storage.path(filename)
        docx_path = default_storage.path(filename + '.docx')
        cv = Converter(pdf_path)
        cv.convert(docx_path, start=0, end=None)
        cv.close()

        # Read the converted DOCX file
        with default_storage.open(docx_path, 'rb') as f:
            docx_data = f.read()

        response = HttpResponse(docx_data, content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        response['Content-Disposition'] = 'attachment; filename=' + pdf_file.name.split('.')[0] + '.docx'

        # Clean up the temporary files
        default_storage.delete(filename)
        default_storage.delete(docx_path)

        return response

    return render(request, 'Common_Page_Tools/convert_pdf_to_docx.html')

def Common_convert_pdf_to_excel(request):
    if request.method == 'POST' and request.FILES['pdf_file']:
        # get the uploaded PDF file
        pdf_file = request.FILES['pdf_file']
        # convert the PDF file into a list of lists
        tables = tabula.read_pdf(pdf_file, pages='all', multiple_tables=True)
        # convert the list of lists into a DataFrame
        df = pd.DataFrame(tables[0])
        # save the DataFrame as an Excel file
        df.to_excel('output.xlsx', index=False)
        # send the Excel file to the user for download
        with open('output.xlsx', 'rb') as excel_file:
            response = HttpResponse(
                excel_file.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=output.xlsx'
            return response
    else:
        return render(request, 'Common_Page_Tools/convert_pdf_to_excel.html')


def Common_convert_excel_to_pdf(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active
        pdf_buffer = io.BytesIO()
        c = canvas.Canvas(pdf_buffer, pagesize=letter)
        c.setFont('Helvetica', 12)
        x, y = 1 * inch, 10.5 * inch
        for row in ws.iter_rows():
            for cell in row:
                cell_value = cell.value
                if cell_value is None:
                    cell_value = ''
                else:
                    cell_value = str(cell_value)
                c.drawString(x, y, cell_value)
                x += 1 * inch
            x = 1 * inch
            y -= 0.5 * inch
        c.save()
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="output.pdf"'

        return response

    return render(request, 'Common_Page_Tools/convert_excel_to_pdf.html')


def Common_convert_images_to_pdf(images):
    filename = tempfile.mktemp(".pdf")
    c = canvas.Canvas(filename)
    for image in images:
        img = Image.open(image)
        width, height = img.size
        c.setPageSize((width, height))
        c.drawImage(image, 0, 0, width, height)
        c.showPage()
        img.close()
    c.save()
    return filename


def Common_convert_jpg_to_pdf(request):
    if request.method == 'POST':
        files = request.FILES.getlist('file')
        if len(files) == 0:
            return HttpResponse("No images selected")
        temp_dir = tempfile.mkdtemp()
        for f in files:
            with open(os.path.join(temp_dir, f.name), 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk)

        # Convert images to PDF
        pdf_file = Common_convert_images_to_pdf(
            [os.path.join(temp_dir, f.name) for f in files])

        # Serve the PDF file for download
        with open(pdf_file, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename=converted.pdf'
            return response

        # Delete temporary files

    return render(request, 'Common_Page_Tools/convert_jpg_to_pdf.html')


def Common_convert_jpg_to_word(request):
    if request.method == 'POST' and request.FILES['files']:
        # Get the uploaded images
        images = request.FILES.getlist('files')

        # Create a new Word document
        document = Document()

        # Loop through the images and add them to the document
        for img in images:
            # Open the image and convert it to a stream
            image = Image.open(img)
            stream = io.BytesIO()
            image.save(stream, format='png')
            stream.seek(0)

            # Add the image to the document
            document.add_picture(stream, width=Inches(6))

        # Save the document
        filename = 'images.docx'
        document.save(filename)

        # Download the document
        with open(filename, 'rb') as f:
            response = HttpResponse(f.read(
            ), content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            response['Content-Disposition'] = 'attachment; filename={}'.format(
                filename)
            return response

    return render(request, 'Common_Page_Tools/convert_jpg_to_word.html')


def Common_cgpa_calculator(request):
    if request.method == 'POST':
        total_credits = 0
        total_weighted_points = 0
        for i in range(1, 9):  # Assuming a maximum of 8 subjects
            credit_field = 'credit' + str(i)
            grade_field = 'grade' + str(i)
            credits = int(request.POST.get(credit_field, 0))
            grade_points = get_grade_points(request.POST.get(grade_field, ''))
            total_credits += credits
            total_weighted_points += credits * grade_points
        try:
            cgpa = round(total_weighted_points / total_credits, 2)
            context = {'cgpa': cgpa, 'len': [i for i in range(1, 10)]}
        except:
            context = {'cgpa': 'cgpa', 'len': [i for i in range(1, 10)]}
            print("error/...")
        return render(request, 'Common_Page_Tools/cgpa_calculator.html', context)
    else:
        return render(request, 'Common_Page_Tools/cgpa_calculator.html', context)


def Common_get_grade_points(grade):
    if grade == 'S':
        return 10
    elif grade == 'A':
        return 9
    elif grade == 'B':
        return 8
    elif grade == 'C':
        return 7
    elif grade == 'D':
        return 6
    elif grade == 'E':
        return 5
    else:
        return 0


# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>


def Common_get_subject(request):
    if request.method == 'POST':
        num = request.POST.get('number')
        print(type(int(num)))
        return render(request, 'Common_Page_Tools/gpa_calculator.html', {'num_sub': [i for i in range(1, int(num)+1)]})

    return render(request, 'Common_Page_Tools/num_of_sub.html')


def Common_gpa_calculator(request):
    credits = request.POST.getlist('credits')
    grades = request.POST.getlist('grades')

    total_credits = sum(map(int, credits))
    total_grade_points = 0

    for i in range(len(credits)):
        grade_point = 0
        if grades[i] == 'A+':
            grade_point = 4.0
        elif grades[i] == 'A':
            grade_point = 4.0
        elif grades[i] == 'A-':
            grade_point = 3.7
        elif grades[i] == 'B+':
            grade_point = 3.3
        elif grades[i] == 'B':
            grade_point = 3.0
        elif grades[i] == 'B-':
            grade_point = 2.7
        elif grades[i] == 'C+':
            grade_point = 2.3
        elif grades[i] == 'C':
            grade_point = 2.0
        elif grades[i] == 'C-':
            grade_point = 1.7
        elif grades[i] == 'D+':
            grade_point = 1.3
        elif grades[i] == 'D':
            grade_point = 1.0
        elif grades[i] == 'F':
            grade_point = 0.0
        total_grade_points += int(credits[i]) * grade_point
    try:
        gpa = total_grade_points / total_credits
    except:
        gpa = 0.0
    context = {'gpa': round(gpa, 2)}
    return render(request, 'Common_Page_Tools/gpa_calculator.html', context)
# >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>


def Common_handwriting_converter(request):
    if request.method == 'POST':
        # Get input text from form
        input_text = request.POST.get('input_text')

        # Create a filename for the image
        filename = 'handwriting.png'

        # Generate image using Pywhatkit
        kit.text_to_handwriting(input_text, os.path.join(
            settings.MEDIA_ROOT, filename))

        # Open image file
        with open(os.path.join(settings.MEDIA_ROOT, filename), 'rb') as f:
            response = HttpResponse(f.read(), content_type="image/png")
            response['Content-Disposition'] = 'attachment; filename=' + filename
            return response
    else:
        return render(request, 'Common_Page_Tools/handwriting.html')


def Common_keyword_to_image(request):
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        urls = get_image_url(keyword)
        print(keyword, urls)
        return render(request, 'Common_Page_Tools/keyword_to_image.html', {'image_urls': urls})
    return render(request, 'Common_Page_Tools/keyword_to_image.html')

# views.py


def Common_video_meeting(request):
    return render(request, 'Common_Page_Tools/video_meeting.html')


def Common_Common_tool(request):
    return render(request, "Common_Page_Tools/Common_tool.html")
